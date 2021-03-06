# -*- coding: utf-8 -*-
import json
import time,datetime
from xpinyin import Pinyin
from collections import defaultdict
from django.http import JsonResponse, HttpResponse
from django.core import serializers
from django.db.models import Q
import operator
from Config.time_utils import *
from Config.base import MONITOR_STATUS_DIC
from Mainevent.models import *  #Event_Analyze, Event, Figure, Information, Event_Hashtag_Senwords
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from Systemmanage.models import SensitiveWord
from rest_framework.views import APIView
from rest_framework.schemas import ManualSchema
import re
from django.db.models import Sum,Count
from io import BytesIO
import requests
import openpyxl
from django.utils.http import urlquote


class Show_event(APIView):
    """展示事件列表
       输出{‘event_name(事件名称)’: ,’keywords_dict(事件关键词)’: ,’content(事件内容)’: , ‘begin_date(开始日期)’:  ,‘end_date(结束日期)’: },{},{}"""
    def get(self, request):
        #result = Event.objects.values('e_id','event_name','keywords_dict','begin_date','end_date')
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        count = len(Event.objects.filter(~Q(hidden_status=1),cal_status=2))
        #print(count)
        result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        index_list = {}
        jre=[]
        #t1 = time.time()
        if result.exists():
            for item in result:
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                if item['sensitive_figure_ratio'] is None:
                    figure_rat = '-'
                else:
                    figure_rat = '%.2f%%' % float(item['sensitive_figure_ratio']*100)
                if item['sensitive_info_ratio'] is None:
                    info_rat = '-'
                else:
                    info_rat = '%.2f%%' % float(item['sensitive_info_ratio']*100)
                jre.append({"eid":item['e_id'],"event_name":item['event_name'],"keywords_dict":item['keywords_dict'],'monitor_status':MONITOR_STATUS_DIC[item['monitor_status']],\
                            "begin_date":sdate,"end_date":edate,'sensitive_figure_ratio':figure_rat,'sensitive_info_ratio':info_rat,'count':count})
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无事件"},safe=False)


class Show_event_sort(APIView):
    """展示事件列表
       输出{‘event_name(事件名称)’: ,’keywords_dict(事件关键词)’: ,’content(事件内容)’: , ‘begin_date(开始日期)’:  ,‘end_date(结束日期)’: },{},{}"""
    def get(self, request):
        #result = Event.objects.values('e_id','event_name','keywords_dict','begin_date','end_date')
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        count = len(Event.objects.filter(~Q(hidden_status=1),cal_status=2))
        name = request.GET.get('label')
        order = request.GET.get('order')
        if name == "sensitive_figure_ratio":
            if order == "descending":
                result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-sensitive_figure_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            else:
                result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('sensitive_figure_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        else:
            if order == "descending":
                result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-sensitive_info_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            else:
                result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('sensitive_info_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        index_list = {}
        jre=[]
        #t1 = time.time()
        if result.exists():
            for item in result:
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                if item['sensitive_figure_ratio'] is None:
                    figure_rat = '-'
                else:
                    figure_rat = '%.2f%%' % float(item['sensitive_figure_ratio']*100)
                if item['sensitive_info_ratio'] is None:
                    info_rat = '-'
                else:
                    info_rat = '%.2f%%' % float(item['sensitive_info_ratio']*100)
                jre.append({"eid":item['e_id'],"event_name":item['event_name'],"keywords_dict":item['keywords_dict'],'monitor_status':MONITOR_STATUS_DIC[item['monitor_status']],\
                            "begin_date":sdate,"end_date":edate,'sensitive_figure_ratio':figure_rat,'sensitive_info_ratio':info_rat,'count':count})
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无事件"},safe=False)



class Show_event_info(APIView):
    """展示事件详情,点击事件传入事件eid
       输出{‘event_name(事件名称)’: ,’keywords_dict(事件关键词)’: ,’content(事件内容)’: , ‘begin_date(开始日期)’:  ,‘end_date(结束日期)’: },{},{}"""
    def get(self, request):
        eid = request.GET.get("eid")
        result = Event.objects.filter(e_id =eid).values('event_name','keywords_dict','begin_date','end_date','monitor_status')
        #item = Event.objects.filter(e_id =eid).values('event_name','keywords_dict','begin_date','end_date').first()
        
        #weibo_count = 0
        #user_count = 0
        jre = {}
        if result.exists():
        #if len(item):
            for item in result:
                weibo_count = 0
                user_count = 0
                #figure_count = 0
                info_count = 0 
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                #eid = item['e_id']
                #e_re = Event.objects.filter(e_id =eid).first()
                #figure_count = len(e_re.figure.all())
                #print(figure_count)
                #info_count = len(e_re.information.all())
                #print(info_count)
                all_re = Event_Analyze.objects.filter(event_name =eid).values('weibo_count','user_count','info_count','figure_count')
                if all_re.exists():
                    #weibo_count = all_re[0]['weibo_count']
                    #user_count = all_re[0]['user_count']
                    for re in all_re:
                        weibo_count += int(re['weibo_count'])
                        user_count += int(re['user_count'])
                        info_count += int(re['info_count'])
                    figure_rat = '0%'
                    info_rat = '0%'
                    if user_count is None:
                        figure_rat = '-'
                    if user_count != 0:
                        figure_rat = '%.2f%%' % float(all_re[0]['figure_count']/user_count * 100)
                    if weibo_count is None:
                        info_rat = '-'
                    if weibo_count != 0:
                        info_rat = '%.2f%%' % float(info_count/weibo_count*100)
                    jre["event_name"]=item['event_name']
                    jre["keywords_dict"]=item['keywords_dict']
                    jre["begin_date"] = sdate
                    jre['end_date'] = edate
                    jre['user_count'] = user_count
                    jre['weibo_count'] = weibo_count
                    jre['sensitive_figure_ratio'] = figure_rat
                    jre['sensitive_info_ratio']=info_rat
                    jre['monitor_status'] = MONITOR_STATUS_DIC[item['monitor_status']]
                else:
                    jre["event_name"]=item['event_name']
                    jre["keywords_dict"]=item['keywords_dict']
                    jre["begin_date"] = sdate
                    jre['end_date'] = edate
            #re = json.dumps(jre,ensure_ascii=False)
            #re = json.loads(re)
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无事件详情"},safe=False)



class search_event(APIView):
    """搜索事件 输入事件标题title 输出'event_name','keywords_dict','content','begin_date','end_date'"""
    def get(self, request):
        jre = []
        name = request.GET.get("title")
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = name) | Q(keywords_dict__contains=name),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        #count = Event.objects.filter(event_name__contains = name).aggregate(count = Count('e_id'))['count']
        count = len(Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = name) | Q(keywords_dict__contains=name),cal_status=2))
        #print(len(result))
        if result.exists():
            for item in result:
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                if item['sensitive_figure_ratio'] is None:
                    figure_rat = '-'
                else:
                    figure_rat = '%.2f%%' % float(item['sensitive_figure_ratio']*100)
                if item['sensitive_info_ratio'] is None:
                    info_rat = '-'
                else:
                    info_rat = '%.2f%%' % float(item['sensitive_info_ratio']*100)
                jre.append({"eid":item['e_id'],"event_name":item['event_name'],"keywords_dict":item['keywords_dict'],'monitor_status':MONITOR_STATUS_DIC[item['monitor_status']],\
                            "begin_date":sdate,"end_date":edate,'sensitive_figure_ratio':figure_rat,'sensitive_info_ratio':info_rat,'count':count})
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无事件"},safe=False)



class search_event_sort(APIView):
    """搜索事件 输入事件标题title 输出'event_name','keywords_dict','content','begin_date','end_date'"""
    def get(self, request):
        jre = []
        info = request.GET.get("title")
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        name = request.GET.get('label')
        order = request.GET.get('order')
        count = len(Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = info) | Q(keywords_dict__contains=info),cal_status=2))
        #print(len(result))
        if name == "sensitive_figure_ratio":
            if order == "descending":
                result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = info) | Q(keywords_dict__contains=info),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-sensitive_figure_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            else:
                result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = info) | Q(keywords_dict__contains=info),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('sensitive_figure_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        else:
            if order == "descending":
                result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = info) | Q(keywords_dict__contains=info),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('-sensitive_info_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            else:
                result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = info) | Q(keywords_dict__contains=info),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date','monitor_status','sensitive_figure_ratio','sensitive_info_ratio').order_by('sensitive_info_ratio','-monitor_status','-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        #result = Event.objects.filter(~Q(hidden_status=1),Q(event_name__contains = name) | Q(keywords_dict__contains=name),cal_status=2).values('e_id','event_name','keywords_dict','begin_date','end_date').order_by('-begin_date')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
        #count = Event.objects.filter(event_name__contains = name).aggregate(count = Count('e_id'))['count']
        if result.exists():
            for item in result:
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                if item['sensitive_figure_ratio'] is None:
                    figure_rat = '-'
                else:
                    figure_rat = '%.2f%%' % float(item['sensitive_figure_ratio']*100)
                if item['sensitive_info_ratio'] is None:
                    info_rat = '-'
                else:
                    info_rat = '%.2f%%' % float(item['sensitive_info_ratio']*100)
                jre.append({"eid":item['e_id'],"event_name":item['event_name'],"keywords_dict":item['keywords_dict'],'monitor_status':MONITOR_STATUS_DIC[item['monitor_status']],\
                            "begin_date":sdate,"end_date":edate,'sensitive_figure_ratio':figure_rat,'sensitive_info_ratio':info_rat,'count':count})
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无事件"},safe=False)




class delete_event(APIView):
    """删除事件"""
    def get(self, request):
        eid = request.GET.get("eid")
        result = Event.objects.filter(e_id=eid)
        if result.exists():
            try:
                Event.objects.filter(e_id=eid).delete() 
                return JsonResponse({"status":201, "msg": "事件已删除"},safe=False,json_dumps_params={'ensure_ascii':False})
            except:
                return JsonResponse({"status":400, "error": "删除失败"},safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "事件不存在"},safe=False,json_dumps_params={'ensure_ascii':False})


class show_cal_event(APIView):
    """展示事件计算任务"""
    def get(self, request):
        cal_status_dic = {
            0: "未计算",
            1: "计算中",
            2: "计算完成"
        }
        jre = []
        result = Event.objects.filter(~Q(hidden_status=1),cal_status__in=[0,1]).values('e_id','event_name','begin_date','end_date','cal_status')
        if result.exists():
            for item in result:
                sdate = item['begin_date'].strftime('%Y-%m-%d')
                if item['end_date'] is None:
                    edate = "至今"
                else:
                    edate = item['end_date'].strftime('%Y-%m-%d')
                jre.append({"eid":item["e_id"],"event_name":item['event_name'],"cal_status":cal_status_dic[item['cal_status']],\
                                "begin_date":sdate,"end_date":edate})
            return JsonResponse(jre,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "事件任务不存在"},safe=False,json_dumps_params={'ensure_ascii':False})






class delete_cal_event(APIView):
    """删除事件计算任务"""
    def get(self, request):
        eid = request.GET.get("eid")
        result = Event.objects.filter(e_id=eid)
        if result.exists():
            try:
                Event.objects.filter(e_id=eid).delete() 
                return JsonResponse({"status":201, "msg": "计算任务已删除"},safe=False,json_dumps_params={'ensure_ascii':False})
            except:
                return JsonResponse({"status":400, "error": "删除失败"},safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "事件不存在"},safe=False,json_dumps_params={'ensure_ascii':False})




class Add_event(APIView):
    """
    添加新的事件计算任务
    """
    def get(self, request):
        event_name = request.GET.get("event_name")
        keywords_dict = request.GET.get("keywords_dict")
        sensitive_word_white = request.GET.get("sensitive_word_white")
        sensitive_word_black = request.GET.get("sensitive_word_black")
        begin_date = request.GET.get("begin_date")
        end_date = request.GET.get("end_date")

        if not event_name or not keywords_dict:
            return JsonResponse({"status":400, "info": "添加失败，缺少必填项！"},safe=False)

        if not begin_date:
            begin_date = ts2date(date2ts(today()) - 19 * 86400)

        event_name_pinyin = Pinyin().get_pinyin(event_name, '')
        e_id = "{}_{}".format(event_name_pinyin, str(int(time.time())))
        Event.objects.create(
            e_id=e_id,
            event_name=event_name,
            keywords_dict=keywords_dict,
            begin_date=begin_date,
            end_date=end_date,
            es_index_name=e_id,
            cal_status=0,
            monitor_status=1
        )

        if sensitive_word_white:
            for word_white in sensitive_word_white.split("&"):
                SensitiveWord.objects.create(
                    s_id=word_white + e_id, 
                    prototype=word_white, 
                    e_id=e_id, 
                    perspective_bias=1
                )

        if sensitive_word_black:
            for word_black in sensitive_word_black.split("&"):
                SensitiveWord.objects.create(
                    s_id=word_black + e_id, 
                    prototype=word_black, 
                    e_id=e_id, 
                    perspective_bias=2
                )

        return JsonResponse({"status":200, "info": "添加成功！"},safe=False)


class Event_trend(APIView):
    """事件热度、敏感度、负面指数
       输入事件id:eid
       输出{"hot_index(热度)": ,"sensitive_index(敏感度)": ,"negative_index(负面指数)":}"""
    def get(self, request):
        """获取事件id:eid"""
        event_id = request.GET.get('eid') 
        res_dict= defaultdict(dict)
        hot = []
        sensitive = []
        neg = []
        date = []
        times = time.time()
        result = Event_Analyze.objects.filter(event_name = event_id)
        if result.exists():
            for re in result:
                date.append(re.into_date.strftime('%Y-%m-%d'))
                hot.append(re.hot_index)
                sensitive.append(re.sensitive_index) 
                neg.append(re.negative_index)
                #print(hot)
            res_dict["date"] = date
            res_dict["hot_index"] = hot
            res_dict["sensitive_index"] = sensitive
            res_dict["negative_index"] = neg
                #res_dict.append({"hot_index":re.hot_index,"sensitive_index":re.sensitive_index,"negative_index":re.negative_index})
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "事件不存在"},safe=False,json_dumps_params={'ensure_ascii':False})


class representative_info(APIView):
    """输入事件id:eid 
       输出代表微博列表：[{"uid": ,"comment": ,"retweeted": ,"date": ,"text": ,"hazard": },{}...]"""
    def get(self, request):
        event_id = request.GET.get('eid')
        etime = request.GET.get('time')
        td1 = etime + " 00:00:00"
        ta1 = time.strptime(td1, "%Y-%m-%d %H:%M:%S")
        ts1 = int(time.mktime(ta1))
        td2 = etime + " 23:59:59"
        ta2 = time.strptime(td2, "%Y-%m-%d %H:%M:%S")
        ts2 = int(time.mktime(ta2))
        res = []
        e = Event.objects.filter(e_id = event_id)
        for item in e:
            info = item.information.all().filter(timestamp__range=(ts1,ts2),cal_status=2).order_by("-hazard_index")[:5]
            for i in info:
                lt = time.localtime(i.timestamp)
                itime = time.strftime("%Y-%m-%d %H:%M:%S",lt)
                if i.nick_name is None:
                    nick = i.uid
                else:
                    nick = i.nick_name
                res.append({"uid":nick,"geo":i.geo,"date":itime,"text":i.text,"hazard":'%.0f' % i.hazard_index})
        return JsonResponse(res,safe=False,json_dumps_params={'ensure_ascii':False})



class related_figure(APIView):
    """事件-人物关联分析"""

    def get(self, request):
        """获取事件eid，返回该事件的相关人物和相关信息,
        数据格式{
                    "figure":[{"f_id": uid, "nick_name": nick_name,"fansnum":fansnum,"friendsnum":friendsnum,"domian":domain,"political":political},{}],
                    "information":[{"text": text1, "hazard_index": hazard_index1},{}]
                }
        """
        res_dict = defaultdict(list)
        eid = request.GET.get('eid')
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        res_event = Event.objects.filter(e_id=eid)  #.first().event_set.all()
        if res_event.exists():
            #print(e)
            res = res_event[0].figure.filter(computestatus=2,identitystatus=1).order_by('-info_count','-event_count')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            count = res_event[0].figure.filter(computestatus=2,identitystatus=1).count()
            res_dict['count'] = count
            for f in res:
                res_dict['table'].append(["f_id","nick_name","info_count","event_count"])
                
                f_id = f.f_id if f.f_id else "未知"
                nick_name = f.nick_name if f.nick_name else "未知"
                info_count = f.info_count if f.info_count else "未知"
                event_count = f.event_count if f.event_count else "未知"
                res_dict['data'].append([f_id, nick_name,info_count,event_count])
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "无相关人物"},safe=False)
        '''
        if len(res_dict):
            page = Paginator(res_dict, limit)
            #page_id = request.GET.get('page_id')
            if page_id:
                try:
                    results = page.page(page_id)
                except PageNotAnInteger:
                    results = page.page(1)
                except EmptyPage:
                    results = page.page(1)
            else:
                results = page.page(1)
            re = json.dumps(list(results),ensure_ascii=False)
            re = json.loads(re)
            '''



class related_info(APIView):
    """事件-信息关联分析"""

    def get(self, request):
        """获取事件eid，返回该事件的相关人物和相关信息,
        数据格式{
                    "figure":[{"f_id": uid, "nick_name": nick_name,"fansnum":fansnum,"friendsnum":friendsnum,"domian":domain,"political":political},{}],
                    "information":[{"text": text1, "hazard_index": hazard_index1},{}]
                }
        """
        res_dict = defaultdict(list)
        eid = request.GET.get('eid')
        limit = request.GET.get("limit")
        page_id = request.GET.get('page_id')
        if page_id is None:
            page_id = 1
        if limit is None:
            limit = 10
        res_event = Event.objects.filter(e_id=eid)  #.first().event_set.all()
        if res_event.exists():
            #print(e)
            res1 = res_event[0].information.filter(cal_status=2,add_manully=0).order_by('-hazard_index', '-timestamp')[int(limit)*(int(page_id)-1):int(limit)*int(page_id)]
            count = res_event[0].information.filter(cal_status=2,add_manully=0).count()
            res_dict['count'] = count
            #print(count)
            for i in res1:
                lt = time.localtime(i.timestamp)
                itime = time.strftime('%Y-%m-%d %H:%M:%S',lt)
                res_dict['table'].append(['text','time','geo','id','hazard_index'])
                res_dict['data'].append([i.text,itime,i.geo,i.i_id,int(i.hazard_index)])
                #print(res_dict["info"])
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无相关信息"},safe=False)
        '''
        if len(res_dict):
            page = Paginator(res_dict, limit)
            #page_id = request.GET.get('page_id')
            if page_id:
                try:
                    results = page.page(page_id)
                except PageNotAnInteger:
                    results = page.page(1)
                except EmptyPage:
                    results = page.page(1)
            else:
                results = page.page(1)
            re = json.dumps(list(results),ensure_ascii=False)
            re = json.loads(re)
            '''



class event_geo_out(APIView):
    """事件国外地域分布"""
    def get(self,request):
        eid = request.GET.get('eid')
        #eid = request.GET.get('eid')
        result = Event_Analyze.objects.filter(event_name=eid).values('geo_weibo_outland')
        geo_dict={}
        #print(type(result))
        if result.exists():
            for re in result:
                for k,v in re["geo_weibo_outland"].items():
                    if k == "未知":
                        continue
                    else:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v
                geo_dict=dict(sorted(geo_dict.items(),key=lambda x:x[1],reverse=True)[:10])
            return JsonResponse(geo_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无该事件地域分布信息"},safe=False)


class event_geo_in(APIView):
    """事件国内地域分布"""
    def get(self,request):
        eid = request.GET.get('eid')
        result = Event_Analyze.objects.filter(event_name =eid).values('geo_weibo_inland')
        geo_dict={}
        b = '未知'
        #print(type(result))
        if result.exists():
            for re in result:
                for k,v in re["geo_weibo_inland"].items():
                    if k == "未知":
                        continue
                    else:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v
                geo_dict=dict(sorted(geo_dict.items(),key=lambda x:x[1],reverse=True)[:10])
            return JsonResponse(geo_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无该事件地域分布信息"},safe=False)


class info_geo_out(APIView):
    """事件按敏感信息国外地域分布"""
    def get(self,request):
        eid = request.GET.get('eid')
        #eid = request.GET.get('eid')
        result = Event_Analyze.objects.filter(event_name=eid).values('geo_info_outland')
        geo_dict={}
        #print(type(result))
        if result.exists():
            #json_data = serializers.serialize("json",result1)
            #results = json.loads(json_data)
            for re in result:
                for k,v in re["geo_info_outland"].items():
                    if k == "未知":
                        continue
                    else:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v

                geo_dict=dict(sorted(geo_dict.items(),key=lambda x:x[1],reverse=True)[:10])
            return JsonResponse(geo_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无该事件地域分布信息"},safe=False)


class info_geo_in(APIView):
    """事件国内按敏感信息地域分布"""
    def get(self,request):
        eid = request.GET.get('eid')
        result = Event_Analyze.objects.filter(event_name =eid).values('geo_info_inland')
        geo_dict={}
        #print(type(result))
        if result.exists():
            for re in result:
                for k,v in re["geo_info_inland"].items():
                    if k == "未知":
                        continue
                    else:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v
                geo_dict=dict(sorted(geo_dict.items(),key=lambda x:x[1],reverse=True)[:10])
            return JsonResponse(geo_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无该事件地域敏感信息分布信息"},safe=False)



class geo_info(APIView):
    """事件地域代表敏感信息分布"""
    def get(self,request):
        geo = request.GET.get('geo')
        eid = request.GET.get('eid')
        info_dict=[]
        result = Event.objects.filter(e_id=eid).first().information.all().filter(geo__contains=geo,cal_status=2).order_by("-hazard_index")[:5]
        if result.exists():
            for i in result:
                lt = time.localtime(i.timestamp)
                itime = time.strftime('%Y-%m-%d %H:%M:%S',lt)
                if i.nick_name is None:
                    nick = i.uid
                else:
                    nick = i.nick_name
                info_dict.append({"uid":nick,"text":i.text,"geo":i.geo,"time":itime,"hazard":'%.0f' % i.hazard_index})
            return JsonResponse(info_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无敏感信息"},safe=False)


class semantic_tl(APIView):
    """事件语义分析之时间轴"""
    def get(self,request):
        eid = request.GET.get('eid')
        result = Event.objects.filter(e_id=eid).first().information.all().filter(~Q(message_type=3),cal_status=2).order_by("timestamp","-hazard_index")    #.order_by()
        #print(result)
        info_dict = defaultdict(list)
        tl_dict ={}
        timeline =[]
        if result.exists():
            '''
            for re in result:
                date = time.strftime('%Y-%m-%d',time.localtime(re[0].timestamp))
                timeline["date"].append(date)
                timeline["content"].append({"source":re[0].uid,"text":re[0].text,"hazard_index":re[0].hazard_index})
            '''
            for re in result:
                #fresult = Event.objects.filter(e_id=eid).first().figure.all().filter(f_id=re.uid).values('nick_name')
                #if fresult.exists():
                    #nick = fresult[0]['nick_name']
                #else:
                    #nick = re.uid
                if re.nick_name is None:
                    nick = re.uid
                else:
                    nick = re.nick_name
                date = time.strftime('%Y-%m-%d',time.localtime(re.timestamp))
                info_dict[date].append({"日期":date,"发博用户昵称":nick,"微博内容":re.text,"危害指数":'%.0f' % re.hazard_index})
                #info_dict[date].append({"date":date,"source":re.uid,"text":re.text,"hazard_index":re.hazard_index})
            for k,v in info_dict.items():
                #timeline.append(sorted(v,key=operator.itemgetter('危害指数'),reverse=True))
                timeline.append(sorted(v,key=operator.itemgetter('危害指数'),reverse=True)[:1][0])
            #print(timeline)
            return JsonResponse(timeline,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无敏感信息"},safe=False)



class semantic_topic(APIView):
    """事件语义分析之主题词"""
    def get(self,request):
        eid = request.GET.get('eid')
        result = Event_Semantic.objects.filter(e_id=eid).values("topics")
        topics = defaultdict(dict)
        the = []
        #the=defaultdict(list)
        if result.exists():
            #print(result)
            #for re in result:
                #print(re["topics"]['0'])
            for i in range(5):
                t=[]
                themes = defaultdict(list)
                for k,v in result[0]["topics"][str(i)].items():
                    #dict(zip(re["topics"][i]['主题'],re["topics"][i]['概率']))
                    try:
                        topics[str(i)][k] += float(v)
                    except:
                        topics[str(i)][k] = float(v)
                topic = dict(sorted(topics[str(i)].items(),key=lambda x:x[1],reverse=True)[:10])
                    #print(topic)
                    #break
                for k,v in topic.items():
                        #print(k,v)
                        t.append({"value":v,"absName":'2',"name":k,"children":[]})
                themes['value'] = 3
                themes['absName'] = 'root'
                themes['name'] = "主题%s" % str(i)
                themes['children'] = t
                
            #print(themes)
                the.append(themes)
            return JsonResponse(the,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无主题信息"},safe=False)



class semantic_info(APIView):
    """事件语义分析之主题词"""
    def get(self,request):
        eid = request.GET.get('eid')
        topic = request.GET.get('topic')
        #topic = topic.encode("unicode_escape")
        #pattern = re.compile(r'topic')
        result = Event.objects.filter(e_id=eid).first().information.all().filter(text__contains=topic,cal_status=2).order_by("-hazard_index")[:5]
        #result = Event.objects.filter(e_id=eid).first().information.all()
        #print(result)
        
        info_dict = []
        if result.exists():
            for res in result:
                itime = time.strftime('%Y-%m-%d %H-%M-%S',time.localtime(res.timestamp))
                if res.nick_name is None:
                    nick = res.uid
                else:
                    nick = res.nick_name
                info_dict.append({"text":res.text,"time":itime,"uid":nick,"geo":res.geo,"hazard":'%.0f' % res.hazard_index})
                #k = pattern.search(res.text)
                #if k:
                    #itime = time.strftime('%Y-%m-%d %H-%M-%S',time.localtime(res.timestamp))
                    #info_dict.append({"text":res.text,"time":itime,"uid":res.uid,"geo":res.geo,"hazard":hazard_index})
            #info_dict = sorted(info_dict,key=operator.itemgetter('hazard'),reverse=True)[:5]
            return JsonResponse(info_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            return JsonResponse({"status":400, "error": "无敏感信息"},safe=False)
           








class Person_show(APIView):
    """用户主表展示"""
    def get(self, request):
        """
        用户展示大表信息
        输入：
        pagenum：当前页码数（以1起始）
        pagelimit：每页显示个数
        输出：
        uid：用户id
        nick_name：人物昵称
        create_at：注册时间
        user_location：注册地点
        fansnum：粉丝数
        friendsnum：关注数
        information_num：发布敏感信息数
        events_name：相关渗透事件
        """
        page_num = int(request.GET.get('pagenum', 1)) - 1
        page_limit = int(request.GET.get('pagelimit', 10))
        persons = Figure.objects.all()[page_num * page_limit : (page_num + 1) * page_limit]
        results = []

        # 批量得到用户相关信息统计
        persons_uid = [p.uid for p in persons]
        informations = Information.objects.filter(uid__in=persons_uid)
        information_count = {uid:0 for uid in persons_uid}
        for i in informations:
            information_count[i.uid] += 1

        # 批量得到用户相关事件统计
        events = Event.objects.filter(figure__uid__in=persons_uid).values_list("event_name", "figure__uid")
        events_name = {uid:[] for uid in persons_uid}
        for e in events:
            event_name, uid = e
            events_name[uid].append(event_name)

        for person in persons:
            dic = {
                "uid":person.uid,
                "nick_name":person.nick_name,
                "create_at":ts2date(int(person.create_at)),
                "user_location":person.user_location,
                "fansnum":person.fansnum,
                "friendsnum":person.friendsnum,
                "information_num":information_count[person.uid],
                "events_name":events_name[person.uid]
            }
            results.append(dic)
        return JsonResponse(results, safe=False)

class Event_Group(APIView):
    def get(self, request):
        e_id = request.GET.get('e_id')

        results = Event_Hashtag_Senwords.objects.filter(e_id=e_id, show_status=1)

        result = {
            'hashtag':{},
            'global_senword':{},
            'event_senword':{}
        }
        if results.exists():
            hashtag = results[0].hashtag
            global_senword = results[0].global_senword
            event_senword = results[0].event_senword

            hashtag = sorted(hashtag.items(), key=lambda x:x[1], reverse=True)[:10]
            global_senword = sorted(global_senword.items(), key=lambda x:x[1], reverse=True)[:10]
            event_senword = sorted(event_senword.items(), key=lambda x:x[1], reverse=True)[:10]

            hashtag = [{"name": item[0], "value": item[1]} for item in hashtag]
            global_senword = {item[0]: item[1] for item in global_senword}
            event_senword = {item[0]: item[1] for item in event_senword}

            result['hashtag'] = hashtag
            result['global_senword'] = global_senword
            result['event_senword'] = event_senword

        return JsonResponse(result, safe=False)



class first_info_trend(APIView):
    """事件敏感度
       输入事件id:eid
       输出{"日期1": ,"日期2":,...}"""
    def get(self, request):
        """获取事件id:eid"""
        event_id = request.GET.get('eid') 
        res_dict= defaultdict(list)
        hot = {}
        sensitive = {}
        neg = {}
        times = time.time()
        res = Event.objects.filter(e_id = event_id)
        if res.exists():
            sd = res[0].begin_date
            ed = res[0].end_date
            if sd and ed:
                result = Event_Analyze.objects.filter(event_name = event_id,into_date__gte=sd,into_date__lte=ed)
            else:
                result = Event_Analyze.objects.filter(event_name = event_id)
            if result.exists():
                for re in result:
                    #res_dict["date"].append(re.into_date.strftime('%Y-%m-%d'))
                    date = re.into_date.strftime('%Y-%m-%d')
                #hot[date] = re.hot_index
                    sensitive[date] = re.sensitive_index
                    #res_dict["sensitive"].append()
                #neg[date] = re.negative_index
                #print(hot)
                #res_dict.append({"hot_index":re.hot_index,"sensitive_index":re.sensitive_index,"negative_index":re.negative_index})
                return JsonResponse(sensitive,safe=False,json_dumps_params={'ensure_ascii':False})
            else:
                return JsonResponse({"status":400, "error": "敏感度待计算"},safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "请输入正确的事件ID"},safe=False,json_dumps_params={'ensure_ascii':False})

import re
class first_info_geo(APIView):
    """事件敏感信息地域分布"""
    def get(self,request):
        eid = request.GET.get('eid')
        result = Event_Analyze.objects.filter(event_name =eid).values('geo_info_inland','geo_info_outland')
        geo_dict={}
        #print(type(result))
        if result.exists():
            pattern=re.compile(r'(\u672a\u77e5)')
            for res in result:
                for k,v in res["geo_info_inland"].items():
                    if pattern.match(k) is None:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v
                for k,v in res["geo_info_outland"].items():
                    if pattern.match(k) is None:
                        try:
                            geo_dict[k] += v
                        except:
                            geo_dict[k] = v
                geo_dict=dict(sorted(geo_dict.items(),key=lambda x:x[1],reverse=True)[:10])
            return JsonResponse(geo_dict,safe=False,json_dumps_params={'ensure_ascii':False}) #
        else:
            geo=[]
            return JsonResponse(geo,safe=False)


class first_sensitive(APIView):
    """首页敏感词"""
    def get(self, request):
        eid = request.GET.get('eid')
        results = Event_Hashtag_Senwords.objects.filter(e_id=eid, show_status=1)
        global_senwords = []
        if results.exists():
            #hashtag = results[0].hashtag
            #print(results[0].global_senword)
            global_senword = results[0].global_senword
            #event_senword = results[0].event_senword

            #hashtag = sorted(hashtag.items(), key=lambda x:x[1], reverse=True)[:100]
            global_senword = sorted(global_senword.items(), key=lambda x:x[1], reverse=True)[:50]
            #event_senword = sorted(event_senword.items(), key=lambda x:x[1], reverse=True)[:10]
            #print(global_senword)
            #hashtag = {item[0]: item[1] for item in hashtag}
            for item in global_senword:
                #print(item[0])
                global_senwords.append({"name":item[0],"value": item[1]})
            #event_senword = {item[0]: item[1] for item in event_senword}

        return JsonResponse(global_senwords, safe=False)


class first_figure(APIView):
    def get(self, request):
        eid = request.GET.get('eid')
        res_dict = []
        res_event = Event.objects.filter(e_id=eid)  #.first().event_set.all()
        if res_event.exists():
            res = res_event[0].information.all().values("uid").annotate(info_count = Count('i_id')).order_by('-info_count')[:10].values("uid","info_count")
            '''
            #res = res_event[0].information.all().values("uid").annotate(info_count = Count('i_id')).order_by('-info_count')[:10].values_list("uid",flat = True)
            res = list(res)
            res_figure = Figure.objects.filter(f_id__in = res).values('nick_name')
            #print(res_figure)
            re = json.dumps(list(res_figure),ensure_ascii=False)
            re = json.loads(re)
            '''
            for re in res:
                result = Figure.objects.filter(f_id = re["uid"])
                #print(result)
                if result.exists():
                    if result[0].nick_name is None:
                        nick = re["uid"]
                    else:
                        nick = result[0].nick_name
                    res_dict.append({"nick_name":nick,"info_count":re["info_count"],"uid":re["uid"]})
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "未找到该事件敏感人物"},safe=False)



class first_info(APIView):
    def get(self, request):
        res_dict = []
        eid = request.GET.get('eid')
        res_event = Event.objects.filter(e_id=eid)  #.first().event_set.all()
        if res_event.exists():
            #print(e)
            res = res_event[0].information.all().filter(cal_status=2).order_by('-hazard_index')[:5]
            for i in res:
                lt = time.localtime(i.timestamp)
                itime = time.strftime('%Y-%m-%d %H:%M:%S',lt)
                if i.nick_name is None:
                    nick = i.uid
                else:
                    nick = i.nick_name
                res_dict.append({"uid":nick,"time":itime,"text":i.text,"i_id":i.i_id})
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "未找到该事件敏感信息"},safe=False)


class first_event(APIView):
    def get(self, request):
        res_dict = []
        result = Event.objects.filter(~Q(hidden_status=1),cal_status=2).order_by('-end_date')[:12]
        if result.exists():
            for res in result:
                '''
                eid = res.e_id
                re = Event.objects.filter(e_id=eid)[0].information.all().order_by('-hazard_index')
                if re.exists():
                    #print(re)
                    #res_dict.append({"eid":res.e_id,"begin_date":res.begin_date,"end_date":res.end_date,"event_name":res.event_name,"text":re[0].text})
                else:
                    res_dict.append({"eid":res.e_id,"begin_date":res.begin_date,"end_date":res.end_date,"event_name":res.event_name,"text":None})
            '''
                res_dict.append({"eid":res.e_id,"begin_date":res.begin_date,"end_date":res.end_date,"event_name":res.event_name,"keywords":res.keywords_dict})
            return JsonResponse(res_dict,safe=False,json_dumps_params={'ensure_ascii':False})
        else:
            return JsonResponse({"status":400, "error": "没有事件信息"},safe=False)





class create_time(APIView):
    """群体账号创建时间展示"""
    def get(self,request):
        e_id = request.GET.get('e_id')
        result = EventAnalysisShow.objects.filter(e_id =e_id).values('create_time')
        if result.exists():
            result = result[0]['create_time']
        else:
            result = {}
        # print(result)
        # print(type(result))
        return JsonResponse(result)


class age(APIView):
    """群体年龄分布展示"""
    def get(self,request):
        e_id = request.GET.get('e_id')
        result = EventAnalysisShow.objects.filter(e_id =e_id).values('age')
        if result.exists():
            result = result[0]['age']
        else:
            result = {}
        # print(result)
        # print(type(result))
        return JsonResponse(result)


class group_geo(APIView):
    """群体地理位置分布展示"""
    def get(self,request):
        e_id = request.GET.get('e_id')
        result = EventAnalysisShow.objects.filter(e_id =e_id).values('geo')[0]['geo']
        # create_time_dict={}
        print(result)
        print(type(result))
        return JsonResponse(result)


class funs_num(APIView):
    """群体粉丝数分布展示"""
    def get(self,request):
        e_id = request.GET.get('e_id')
        result = EventAnalysisShow.objects.filter(e_id =e_id).values('funs_num')
        if result.exists():
            result = result[0]['funs_num']
        else:
            result = {}
        # print(result)
        # print(type(result))
        return JsonResponse(result)


class friends_num(APIView):
    """群体粉丝数分布展示"""
    def get(self,request):
        e_id = request.GET.get('e_id')
        result = EventAnalysisShow.objects.filter(e_id =e_id).values('friends_num')
        if result.exists():
            result = result[0]['friends_num']
        else:
            result = {}
        # print(result)
        # print(type(result))
        return JsonResponse(result)

class ProcessEvent:
    def __init__(self, event_id):
        self.event_id = event_id
        self.workbook = openpyxl.Workbook()
        self.count = 0
        self.ip = "http://219.224.135.12:9728"

    # 基本信息
    def get_information(self):
        result = []
        title = ["事件ID", "事件名称", "事件关键词", "事件监测状态", "事件开始日期", "事件结束日期", "事件相关敏感人物占事件相关总人物比例", "事件相关敏感信息占事件总微博比例",
                 "事件总微博数", "事件参与总用户数"]
        result.append(title)
        request_url = self.ip +"/Mainevent/einfoshow/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            # outcome["monitor_status"]
            result.append(
                [self.event_id, outcome["event_name"], outcome["keywords_dict"], "test", outcome["begin_date"],
                 outcome["end_date"], outcome["sensitive_figure_ratio"], outcome["sensitive_info_ratio"],
                 outcome["weibo_count"], outcome["user_count"]])
        else:
            return "error"
        self.write_excel_xlsx("基本信息", result)
        return "success"

    # 事件态势
    def get_situation(self):
        result = []
        title = ["日期", "热度", "敏感指数", "负面指数"]
        result.append(title)
        request_url = self.ip +"/Mainevent/trend/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for i in range(len(outcome["date"])):
                result.append([outcome["date"][i], outcome["hot_index"][i], outcome["sensitive_index"][i],
                               outcome["negative_index"][i]])
        else:
            return "error"
        self.write_excel_xlsx("事件态势", result)
        return "success"

    # 地域分析(no?)
    def get_spatial(self):
        result = []
        title = ["国内/国外", "全部信息/敏感信息", "城市", "统计数量"]
        result.append(title)
        # 国内全部信息
        request_url = self.ip +"/Mainevent/geo_in/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for k, v in outcome.items():
                result.append(["国内", "全部信息", k, v])
        else:
            return "error"

        # 国外全部信息
        request_url = self.ip +"/Mainevent/geo_out/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for k, v in outcome.items():
                result.append(["国外", "全部信息", k, v])
        else:
            return "error"

        # 国内敏感信息
        request_url = self.ip +"/Mainevent/info_geo_in/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for k, v in outcome.items():
                result.append(["国内", "敏感信息", k, v])
        else:
            return "error"

        # 国外敏感信息
        request_url = self.ip +"/Mainevent/info_geo_out/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for k, v in outcome.items():
                result.append(["国外", "敏感信息", k, v])
        else:
            return "error"

        self.write_excel_xlsx("地域分析", result)
        return "success"

    # 子事件分析
    def get_childevents(self):
        result = []
        title = ["日期", "代表敏感信息发布者", "代表敏感信息内容", "危害指数"]
        result.append(title)
        request_url = self.ip +"/Mainevent/timeline/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for item in outcome:
                result.append([item["日期"], item["发博用户昵称"], item["微博内容"], item["危害指数"]])
        else:
            return "error"
        self.write_excel_xlsx("子事件分析", result)
        return "success"

    # 主题分析
    def get_toipcs(self):
        result = []
        title = ["主题", "主题词", "概率"]
        result.append(title)
        request_url = self.ip +"/Mainevent/event_topic/?eid={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for item in outcome:
                result.append([item["name"], item["name"], item["value"]])
                for term in item["children"]:
                    result.append([item["name"], term["name"], term["value"]])
        else:
            return "error"
        self.write_excel_xlsx("主题分析", result)
        return "success"

    # 群体分析
    def get_groups(self):
        result = []
        title = ["hashtag/敏感词", "所属分析", "统计数量"]
        result.append(title)
        request_url = self.ip +"/Mainevent/event_group/?e_id={}".format(self.event_id)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for item in outcome["hashtag"]:
                result.append([item["name"], "hashtag分析", item["value"]])
            for k, v in outcome["global_senword"].items():
                result.append([k, "通用敏感词分析", v])
            for k, v in outcome["event_senword"].items():
                result.append([k, "事件敏感词分析", v])
        else:
            return "error"
        self.write_excel_xlsx("群体分析", result)
        return "success"

    # 关联敏感信息
    def get_sensitive(self):
        result = []
        title = ["敏感信息ID", "敏感信息内容", "敏感信息发布地址", "敏感信息发布时间", "危害指数"]
        result.append(title)
        count_url = self.ip +"/Mainevent/related_info/?eid={}&limit=1&page_id=1".format(self.event_id)
        r_count = requests.get(count_url)
        if r_count.status_code == 200:
            count = json.loads(r_count.text)["count"]
        else:
            return "error"
        request_url = self.ip +"/Mainevent/related_info/?eid={}&limit={}&page_id=1".format(
            self.event_id, count)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for item in outcome["data"]:
                result.append([item[3], item[0], item[2], item[1], item[4]])
        else:
            return "error"
        self.write_excel_xlsx("关联敏感信息", result)
        return "success"

    # 关联敏感人物
    def get_sensitive_person(self):
        result = []
        title = ["人物ID", "人物昵称", "人物参与事件数", "人物发布敏感信息数"]
        result.append(title)
        count_url = self.ip +"/Mainevent/related_figure/?eid={}&limit=1&page_id=1".format(
            self.event_id)
        r_count = requests.get(count_url)
        if r_count.status_code == 200:
            count = json.loads(r_count.text)["count"]
        else:
            return "error"
        request_url = self.ip +"/Mainevent/related_figure/?eid={}&limit={}&page_id=1".format(
            self.event_id, count)
        r = requests.get(request_url)
        if r.status_code == 200:
            outcome = json.loads(r.text)
            for item in outcome["data"]:
                result.append([item[0], item[1], item[2], item[3]])
        else:
            return "error"
        self.write_excel_xlsx("关联敏感人物", result)
        return "success"

    # 写入excel
    def write_excel_xlsx(self, sheet_name, value):
        index = len(value)
        if self.count == 0:
            sheet = self.workbook.active
            sheet.title = sheet_name
        else:
            sheet = self.workbook.create_sheet(sheet_name)
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        self.count += 1

    # 传输excel
    def to_excel(self):
        self.get_information()
        self.get_situation()
        self.get_spatial()
        self.get_childevents()
        self.get_toipcs()
        self.get_groups()
        self.get_sensitive()
        self.get_sensitive_person()
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        filename = urlquote(str(self.event_id)+".xlsx")
        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

class event_to_excel(APIView):
    """事件导出excel"""
    def get(self,request):
        eid = request.GET.get('eid')
        event = ProcessEvent(eid)
        result = event.to_excel()
        return result
